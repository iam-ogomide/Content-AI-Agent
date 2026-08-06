"""Calendar exporter — turns a Planner calendar into a downloadable spreadsheet.

Not a sub-agent: it transforms data the Planner already produced, no model
call involved. Output is a plain .xlsx, which opens directly in Google
Sheets (File > Import) or Excel — no Google API credentials required.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SLOT_COLUMNS = ["date", "channel", "pillar", "topic", "angle", "tone", "icp", "cta"]
SLOT_HEADERS = ["Date", "Channel", "Pillar", "Topic", "Angle", "Tone", "ICP", "CTA"]

# Roughly matches how wide each column reads in the frontend's calendar table.
COLUMN_WIDTHS = [12, 14, 22, 30, 36, 12, 30, 26]


def build_calendar_workbook(calendar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Calendar"

    ws.append(SLOT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for slot in calendar.get("slots", []):
        ws.append([slot.get(key, "") for key in SLOT_COLUMNS])

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    notes = [
        ("Coverage notes", calendar.get("coverage_notes")),
        ("Cadence notes", calendar.get("cadence_notes")),
    ]
    if any(text for _, text in notes):
        ws.append([])
        for label, text in notes:
            if text:
                ws.append([label, text])

    return wb


def calendar_filename(calendar):
    timeframe = calendar.get("timeframe") or "plan"
    slug = re.sub(r"[^a-z0-9]+", "-", timeframe.lower()).strip("-") or "plan"
    return f"content-calendar-{slug}.xlsx"


def calendar_to_xlsx_bytes(calendar):
    wb = build_calendar_workbook(calendar)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
